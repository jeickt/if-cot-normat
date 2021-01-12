from openpyxl import load_workbook
import os
import pandas as pd
import re

from Classes import *

listasCursos = []
sequencia = [1, 10, 11, 9, 8, 7, 5, 6, 4, 3, 2]
primeirasChamadas = []


def recuperarListasIniciais():
    arquivos = os.listdir('C:\\temp2')
    for arq in arquivos:
        nome = re.split(r'[-.]+', arq)
        onzeListas = [Curso((nome[0], nome[1]))]
        wb = load_workbook('C:\\temp2\\' + arq)
        for i in range(1, 12, 1):
            listaCota = []
            ws = wb.get_sheet_by_name(f'Cota-{i}')
            verificador = True
            count = 2
            while verificador:
                if ws[f'A{count}'].value == None:
                    verificador = False
                else:
                    listaCota.append(CandResumido(ws[f'A{count}'].value, ws[f'B{count}'].value, ws[f'C{count}'].value,
                                                 None, 0, "SIM"))
                    count += 1
            onzeListas.append(listaCota)

        ws = wb.get_sheet_by_name('Vagas')
        cursoVagas = []
        for i in range(2, 13, 1):
            cursoVagas.append([ws[f'A{i}'].value, ws[f'B{i}'].value])
        onzeListas.append(cursoVagas)

        listasCursos.append(onzeListas)


def montarListaPrimeiraChamada(curso):
    chamada = []

    # Primeira montagem
    for i in range(11):
        if curso[12][i][1] != 0:
            for cand in curso[sequencia[i]]:
                possivelChamado = list(filter(lambda x: x.codigo == cand.codigo, curso[1]))[0]
                if possivelChamado.chamada == 0:
                    possivelChamado.tipoVaga = sequencia[i]
                    possivelChamado.chamada = 1
                    chamada.append(possivelChamado)
                    curso[12][i][1] -= 1
                if curso[12][i][1] == 0:
                    break

    # Remanejamento de vagas
    tipoVaga = 1
    while curso[1][-1].chamada != 1 and any(vagas[1] != 0 for vagas in curso[12]):
        procurandoNovoNomeChamada = True
        if curso[12][tipoVaga][1] > 0:
            for i in range(tipoVaga-1, -1, -1):
                for cand in curso[sequencia[i]]:
                    if list(filter(lambda x: x.codigo == cand.codigo, curso[1])):
                        possivelChamado = list(filter(lambda x: x.codigo == cand.codigo, curso[1]))[0]
                        if possivelChamado.chamada == 0:
                            possivelChamado.tipoVaga = sequencia[tipoVaga]
                            possivelChamado.chamada = 1
                            chamada.append(possivelChamado)
                            curso[12][tipoVaga][1] -= 1
                            procurandoNovoNomeChamada = False
                            break
                if not procurandoNovoNomeChamada:
                    break
        else:
            tipoVaga += 1

    primeirasChamadas.append(chamada)


def fazerArquivosDeChamada():
    for i in range(len(primeirasChamadas)):
        data = []
        for c in range(len(primeirasChamadas[i])):
            data.append([primeirasChamadas[i][c].codigo, primeirasChamadas[i][c].nome, primeirasChamadas[i][c].inscricao,
                         primeirasChamadas[i][c].tipoVaga, primeirasChamadas[i][c].chamada])
        df = pd.DataFrame(data, columns=['Código', 'Nome', 'Inscrição', 'Tipo da Vaga', 'Chamada'])
        df.to_excel(f'C:\\temp-PrimeiraChamada\\PC-{listasCursos[i][0].nome[0]}-{listasCursos[i][0].nome[1]}.xlsx', index=False)


def consolidarConferenciaPrincipal():
    arquivos = os.listdir('C:\\temp2')
    for i in range(len(arquivos)):
        wb = load_workbook('C:\\temp2\\' + arquivos[i])
        ws = wb.get_sheet_by_name('Cota-1')

        for j in range(len(listasCursos[i][1])):
            if any(listasCursos[i][1][j].codigo == cand.codigo for cand in primeirasChamadas[i]):
                ws[f'G{j+2}'] = 1

        wb.save('C:\\temp2\\' + arquivos[i])
        wb.close()
