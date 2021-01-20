from openpyxl import load_workbook
import os
import pandas as pd
import re

from Classes import *

listasCursos = []
listasPrimeiraChamada = []
sequencia = [1, 10, 11, 9, 8, 7, 5, 6, 4, 3, 2]
segundasChamadas = []


def recuperarListasIniciais():
    # recupera os arquivamento de FuncoesIniciais e da Primeira Chamada.
    arquivos = os.listdir('C:\\temp-PrimeiraChamada')
    for arq in arquivos:
        primeiraChamada = []
        wb = load_workbook('C:\\temp-PrimeiraChamada\\' + arq)
        ws = wb.get_sheet_by_name('Sheet1')
        verificador = True
        count = 2
        while verificador:
            if not ws[f'A{count}'].value:
                verificador = False
            else:
                primeiraChamada.append(CandResumido(ws[f'A{count}'].value, ws[f'B{count}'].value, ws[f'C{count}'].value,
                                                    ws[f'D{count}'].value, ws[f'E{count}'].value, None))
                count += 1
        listasPrimeiraChamada.append(primeiraChamada)

    arquivos = os.listdir('C:\\temp2')
    for i in range(len(arquivos)):
        nome = re.split(r'[-.]+', arquivos[i])
        onzeListas = [Curso((nome[0], nome[1]))]
        wb = load_workbook('C:\\temp2\\' + arquivos[i])
        for j in range(1, 12, 1):
            listaCota = []
            ws = wb.get_sheet_by_name(f'Cota-{j}')
            wsVagas = wb.get_sheet_by_name('Vagas')
            verificador = True
            count = 2
            while verificador:
                if not ws[f'A{count}'].value:
                    verificador = False
                else:
                    if j == 1:
                        listaCota.append(CandControle(ws[f'A{count}'].value, ws[f'B{count}'].value,
                                                      ws[f'C{count}'].value, ws[f'D{count}'].value,
                                                      ws[f'E{count}'].value, ws[f'F{count}'].value,
                                                     ws[f'G{count}'].value, ws[f'H{count}'].value,
                                                      ws[f'I{count}'].value, ws[f'J{count}'].value,
                                                      ws[f'K{count}'].value, ws[f'L{count}'].value))
                        if list(filter(lambda x: x.codigo == ws[f'A{count}'].value, listasPrimeiraChamada[i])):
                            cand = list(filter(lambda x: x.codigo == ws[f'A{count}'].value,
                                               listasPrimeiraChamada[i]))[0]
                            if not ws[f'F{count}'].value:
                                wsVagas[f'C{sequencia.index(cand.tipoVaga) + 2}'] = \
                                    wsVagas[f'C{sequencia.index(cand.tipoVaga) + 2}'].value + 1
                                # verifica as vagas que sobraram da chamada anterior.
                            if ws[f'F{count}'].value or ws[f'H{count}'].value:
                                listaCota[-1].valido = "NAO"
                                ws[f'E{count}'] = "NAO"
                                # elimina candidatos que foram matriculados ou completamente desclassificados.
                    else:
                        listaCota.append(CandResumido(ws[f'A{count}'].value, ws[f'B{count}'].value,
                                                      ws[f'C{count}'].value, None, None, ws[f'D{count}'].value))
                    count += 1
            onzeListas.append(listaCota)
        wb.save('C:\\temp2\\' + arquivos[i])

        cursoVagas = []
        for j in range(2, 13, 1):
            cursoVagas.append([wsVagas[f'A{j}'].value, wsVagas[f'C{j}'].value])
            # obtem as vagas remanescentes no curso.

        onzeListas.append(cursoVagas)

        wb.close()
        listasCursos.append(onzeListas)


def verificarDesclassificacoesEmCotas():
    # realiza o tratamento das desclassificações ocorridas nas análises de cada tipo de cota.
    arquivos = os.listdir('C:\\temp2')
    for i in range(len(arquivos)):
        wb = load_workbook('C:\\temp2\\' + arquivos[i])

        for cand in listasCursos[i][1]:
            if cand.descRI:
                for j in [2, 3, 4, 5]:
                    if list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j])):
                        candResumido = list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j]))[0]
                        candResumido.valido = "NAO"
                        ws = wb.get_sheet_by_name(f'Cota-{j}')
                        ws[f'D{listasCursos[i][j].index(candResumido) + 2}'] = "NAO"
            if cand.descPPI:
                for j in [2, 3, 6, 7, 11]:
                    if list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j])):
                        candResumido = list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j]))[0]
                        candResumido.valido = "NAO"
                        ws = wb.get_sheet_by_name(f'Cota-{j}')
                        ws[f'D{listasCursos[i][j].index(candResumido) + 2}'] = "NAO"
            if cand.descEP:
                for j in [2, 3, 4, 5, 6, 7, 8, 9]:
                    if list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j])):
                        candResumido = list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j]))[0]
                        candResumido.valido = "NAO"
                        ws = wb.get_sheet_by_name(f'Cota-{j}')
                        ws[f'D{listasCursos[i][j].index(candResumido) + 2}'] = "NAO"
            if cand.descPCD:
                for j in [2, 4, 6, 8, 10]:
                    if list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j])):
                        candResumido = list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j]))[0]
                        candResumido.valido = "NAO"
                        ws = wb.get_sheet_by_name(f'Cota-{j}')
                        ws[f'D{listasCursos[i][j].index(candResumido) + 2}'] = "NAO"
        wb.save('C:\\temp2\\' + arquivos[i])
        wb.close()


def montarListaSegundaChamada(curso):
    # constrói as listas de Segunda Chamada
    chamada = []

    # Primeira montagem
    for i in range(11):
        if curso[12][i][1] != 0:
            for cand in curso[sequencia[i]]:
                if cand.valido == "SIM":
                    possivelChamado = list(filter(lambda x: x.codigo == cand.codigo, curso[1]))[0]
                    if possivelChamado.valido == "SIM" and possivelChamado.chamada in [0, 1]:
                        possivelChamado.tipoVaga = sequencia[i]
                        possivelChamado.chamada = 2
                        chamada.append(possivelChamado)
                        curso[12][i][1] -= 1
                    if curso[12][i][1] == 0:
                        break

    # Remanejamento de vagas
    tipoVaga = 1
    while any(cand.valido == "SIM" and cand.chamada != 2 for cand in curso[1]) and \
            any(vagas[1] != 0 for vagas in curso[12]):
        procurandoNovoNomeChamada = True
        if curso[12][tipoVaga][1] > 0:
            for i in range(tipoVaga-1, -1, -1):
                for cand in curso[sequencia[i]]:
                    if cand.valido == "SIM":
                        if list(filter(lambda x: x.codigo == cand.codigo, curso[1])):
                            possivelChamado = list(filter(lambda x: x.codigo == cand.codigo, curso[1]))[0]
                            if possivelChamado.valido == "SIM" and possivelChamado.chamada in [0, 1]:
                                possivelChamado.tipoVaga = sequencia[tipoVaga]
                                possivelChamado.chamada = 2
                                chamada.append(possivelChamado)
                                curso[12][tipoVaga][1] -= 1
                                procurandoNovoNomeChamada = False
                                break
                if not procurandoNovoNomeChamada:
                    break
        else:
            tipoVaga += 1

    segundasChamadas.append(chamada)


def fazerArquivosDeChamada():
    # arquiva as listas de Segunda Chamada
    for i in range(len(segundasChamadas)):
        data = []
        for c in range(len(segundasChamadas[i])):
            data.append([segundasChamadas[i][c].codigo, segundasChamadas[i][c].nome, segundasChamadas[i][c].inscricao,
                         segundasChamadas[i][c].tipoVaga, segundasChamadas[i][c].chamada])
        df = pd.DataFrame(data, columns=['Código', 'Nome', 'Inscrição', 'Tipo da Vaga', 'Chamada'])
        df.to_excel(f'C:\\temp-SegundaChamada\\SC-{listasCursos[i][0].nome[0]}-{listasCursos[i][0].nome[1]}.xlsx',
                    index=False)


def consolidarConferenciaPrincipal():
    # Salva no controle principal a informação da chamada para cada candidato de Segunda Chamada.
    arquivos = os.listdir('C:\\temp2')
    for i in range(len(arquivos)):
        wb = load_workbook('C:\\temp2\\' + arquivos[i])
        ws = wb.get_sheet_by_name('Cota-1')

        for j in range(len(listasCursos[i][1])):
            if any(listasCursos[i][1][j].codigo == cand.codigo for cand in segundasChamadas[i]):
                ws[f'G{j+2}'] = 2

        wb.save('C:\\temp2\\' + arquivos[i])
        wb.close()
