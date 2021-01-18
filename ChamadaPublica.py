from openpyxl import load_workbook
import os
import pandas as pd
import re

from Classes import *

listasCursos = []
listasTerceiraChamada = []
sequencia = [1, 10, 11, 9, 8, 7, 5, 6, 4, 3, 2]
chamadasPublicas = []


def recuperarListasIniciais():
    arquivos = os.listdir('C:\\temp-TerceiraChamada')
    for arq in arquivos:
        terceiraChamada = []
        wb = load_workbook('C:\\temp-TerceiraChamada\\' + arq)
        ws = wb.get_sheet_by_name('Sheet1')
        verificador = True
        count = 2
        while verificador:
            if not ws[f'A{count}'].value:
                verificador = False
            else:
                terceiraChamada.append(CandResumido(ws[f'A{count}'].value, ws[f'B{count}'].value, ws[f'C{count}'].value,
                                                    ws[f'D{count}'].value, ws[f'E{count}'].value, None))
                count += 1
        listasTerceiraChamada.append(terceiraChamada)

    arquivos = os.listdir('C:\\temp2')
    for i in range(len(arquivos)):
        nome = re.split(r'[-.]+', arquivos[i])
        onzeListas = [Curso((nome[0], nome[1]))]
        wb = load_workbook('C:\\temp2\\' + arquivos[i])
        for j in range(1, 12, 1):
            listaCota = []
            ws = wb.get_sheet_by_name(f'Cota-{j}')
            wsVagas = wb.get_sheet_by_name('Vagas')
            verificador = True
            count = 2
            while verificador:
                if not ws[f'A{count}'].value:
                    verificador = False
                else:
                    if j == 1:
                        listaCota.append(CandControle(ws[f'A{count}'].value, ws[f'B{count}'].value,
                                                      ws[f'C{count}'].value, ws[f'D{count}'].value,
                                                      ws[f'E{count}'].value, ws[f'F{count}'].value,
                                                      ws[f'G{count}'].value, ws[f'H{count}'].value,
                                                      ws[f'I{count}'].value, ws[f'J{count}'].value,
                                                      ws[f'K{count}'].value, ws[f'L{count}'].value))
                        if list(filter(lambda x: x.codigo == ws[f'A{count}'].value, listasTerceiraChamada[i])):
                            cand = list(filter(lambda x: x.codigo == ws[f'A{count}'].value,
                                               listasTerceiraChamada[i]))[0]
                            if not ws[f'F{count}'].value:
                                wsVagas[f'E{sequencia.index(cand.tipoVaga) + 2}'] = \
                                    wsVagas[f'E{sequencia.index(cand.tipoVaga) + 2}'].value + 1
                            if ws[f'F{count}'].value or ws[f'H{count}'].value:
                                listaCota[-1].valido = "NAO"
                                ws[f'E{count}'] = "NAO"
                    else:
                        listaCota.append(CandResumido(ws[f'A{count}'].value, ws[f'B{count}'].value,
                                                      ws[f'C{count}'].value, None, None, ws[f'D{count}'].value))
                    count += 1
            onzeListas.append(listaCota)
        wb.save('C:\\temp2\\' + arquivos[i])

        cursoVagas = []
        for j in range(2, 13, 1):
            cursoVagas.append([wsVagas[f'A{j}'].value, wsVagas[f'E{j}'].value])

        onzeListas.append(cursoVagas)

        wb.close()
        listasCursos.append(onzeListas)


def verificarDesclassificacoesEmCotas():
    arquivos = os.listdir('C:\\temp2')
    for i in range(len(arquivos)):
        wb = load_workbook('C:\\temp2\\' + arquivos[i])

        for cand in listasCursos[i][1]:
            if cand.descPPI:
                for j in [2, 3, 6, 7, 11]:
                    if list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j])):
                        candResumido = list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j]))[0]
                        candResumido.valido = "NAO"
                        ws = wb.get_sheet_by_name(f'Cota-{j}')
                        ws[f'D{listasCursos[i][j].index(candResumido) + 2}'] = "NAO"
            if cand.descRI:
                for j in [2, 3, 4, 5]:
                    if list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j])):
                        candResumido = list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j]))[0]
                        candResumido.valido = "NAO"
                        ws = wb.get_sheet_by_name(f'Cota-{j}')
                        ws[f'D{listasCursos[i][j].index(candResumido) + 2}'] = "NAO"
            if cand.descEP:
                for j in [2, 3, 4, 5, 6, 7, 8, 9]:
                    if list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j])):
                        candResumido = list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j]))[0]
                        candResumido.valido = "NAO"
                        ws = wb.get_sheet_by_name(f'Cota-{j}')
                        ws[f'D{listasCursos[i][j].index(candResumido) + 2}'] = "NAO"
            if cand.descPCD:
                for j in [2, 4, 6, 8, 10]:
                    if list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j])):
                        candResumido = list(filter(lambda x: x.codigo == cand.codigo, listasCursos[i][j]))[0]
                        candResumido.valido = "NAO"
                        ws = wb.get_sheet_by_name(f'Cota-{j}')
                        ws[f'D{listasCursos[i][j].index(candResumido) + 2}'] = "NAO"
        wb.save('C:\\temp2\\' + arquivos[i])
        wb.close()


def inserirListasVazias():
    for curso in listasCursos:
        vagasTotais = 0
        for vagas in curso[12]:
            vagasTotais += vagas[1]
        chamadasPublicas.append([curso[0], vagasTotais, []])


def comecarChamada():
    opcao = 0
    while True:
        print("Chamar próximo candidato de qual curso?")
        for i in range(len(listasCursos)):
            if chamadasPublicas[i][1] > 0:
                print(f'    Opção {i+1} - Curso {listasCursos[i][0].nome[1]} - Vagas = {chamadasPublicas[i][1]}')
            else:
                print(f'    Opção {i + 1} - Curso {listasCursos[i][0].nome[1]} - Vagas = {chamadasPublicas[i][1]}')
        print(f'Opção 311415118 - para encerrar.')
        try:
            opcao = int(input()) - 1
            if opcao != 311415117 and (opcao < 0 or opcao > len(listasCursos) - 1):
                raise ValueError
        except:
            print("Opção inválida.")
        if opcao == 311415117:
            break
        elif not (opcao < 0 or opcao > len(listasCursos) - 1):
            if chamadasPublicas[opcao][1] <= 0:
                print("Não há mais vagas para este curso")
            else:
                selec = chamarCandidato(listasCursos[opcao], opcao)
                if not selec[0]:
                    print("Não há mais candidatos válidos para este curso.")
                else:
                    chamadasPublicas[opcao][2].append(selec)
                    print(f'{selec[0].codigo} - {selec[0].nome} - Inscrição {selec[0].inscricao} - '
                          f'Vaga {selec[0].tipoVaga} /// vagas restantes no curso = {chamadasPublicas[opcao][1]} ou '
                          f'{chamadasPublicas[opcao][1] - 1}')
                    print("-------------------------------------------------------------------------------------------")



def chamarCandidato(curso, pos):
    if len(chamadasPublicas[pos][2]) > 0:
        verificarMatricula(pos)
    for i in range(11):
        if curso[12][i][1] != 0:
            for cand in curso[sequencia[i]]:
                if cand.valido == "SIM":
                    possivelChamado = list(filter(lambda x: x.codigo == cand.codigo, curso[1]))[0]
                    if possivelChamado.valido == "SIM" and possivelChamado.chamada in [0, 1, 2, 3, "CPn"]:
                        possivelChamado.tipoVaga = sequencia[i]
                        possivelChamado.chamada = "CP"
                        return [possivelChamado,""]

    tipoVaga = 1
    while any(cand.valido == "SIM" for cand in curso[1]) and any(vagas[1] != 0 for vagas in curso[12]):
        if curso[12][tipoVaga][1] > 0:
            for i in range(tipoVaga - 1, -1, -1):
                for cand in curso[sequencia[i]]:
                    if cand.valido == "SIM":
                        if list(filter(lambda x: x.codigo == cand.codigo, curso[1])):
                            possivelChamado = list(filter(lambda x: x.codigo == cand.codigo, curso[1]))[0]
                            if possivelChamado.valido == "SIM" and possivelChamado.chamada in [0, 1, 2, 3, "CPn"]:
                                possivelChamado.tipoVaga = sequencia[tipoVaga]
                                possivelChamado.chamada = "CP"
                                return [possivelChamado, ""]
        else:
            tipoVaga += 1
    return (None, None)


def verificarMatricula(pos):
    candAVerificar = chamadasPublicas[pos][2][-1][0]
    print(f"""O candidato {candAVerificar.nome} realizou a matrícula?
                Opção 1 - Sim;
                Opção 2 - Não. Desclassificado por ausência ou pelos RE;
                Opção 3 - Desclassificado por descRI;
                Opção 4 - Desclassificado por descPPI;
                Opção 5 - Desclassificado por descEP;
                Opção 6 - Desclassificado por descPCD.""")
    opcao = input()
    while opcao not in ["1", "2", "3", "4", "5", "6"]:
        opcao = input(f"Opção inválida. O candidato {candAVerificar.nome} realizou a matrícula? ")

    if opcao == "1":
        chamadasPublicas[pos][2][-1][1] = "matriculado"
        chamadasPublicas[pos][1] -= 1
        listasCursos[pos][12][sequencia.index(candAVerificar.tipoVaga)][1] -= 1
        print("Matrícula confirmada")
    elif opcao == "2":
        chamadasPublicas[pos][2][-1][1] = "não matriculado"
        print("Candidato desclassificado")
    else:
        candAVerificar.chamada = "CPn"
        arquivos = os.listdir('C:\\temp2')
        wb = load_workbook('C:\\temp2\\' + arquivos[pos])
        chamadasPublicas[pos][2][-1][1] = "ainda não matriculado"
        if opcao == "3":
            for i in [2, 3, 4, 5]:
                if list(filter(lambda x: x.codigo == candAVerificar.codigo, listasCursos[pos][i])):
                    candResumido = list(filter(lambda x: x.codigo == candAVerificar.codigo, listasCursos[pos][i]))[0]
                    candResumido.valido = "NAO"
                    ws = wb.get_sheet_by_name(f'Cota-{i}')
                    ws[f'D{listasCursos[pos][i].index(candResumido) + 2}'] = "NAO"
        elif opcao == "4":
            for i in [2, 3, 6, 7, 11]:
                if list(filter(lambda x: x.codigo == candAVerificar.codigo, listasCursos[pos][i])):
                    candResumido = list(filter(lambda x: x.codigo == candAVerificar.codigo, listasCursos[pos][i]))[0]
                    candResumido.valido = "NAO"
                    ws = wb.get_sheet_by_name(f'Cota-{i}')
                    ws[f'D{listasCursos[pos][i].index(candResumido) + 2}'] = "NAO"
        elif opcao == "5":
            for i in [2, 3, 4, 5, 6, 7, 8, 9]:
                if list(filter(lambda x: x.codigo == candAVerificar.codigo, listasCursos[pos][i])):
                    candResumido = list(filter(lambda x: x.codigo == candAVerificar.codigo, listasCursos[pos][i]))[0]
                    candResumido.valido = "NAO"
                    ws = wb.get_sheet_by_name(f'Cota-{i}')
                    ws[f'D{listasCursos[pos][i].index(candResumido) + 2}'] = "NAO"
        elif opcao == "6":
            for i in [2, 4, 6, 8, 10]:
                if list(filter(lambda x: x.codigo == candAVerificar.codigo, listasCursos[pos][i])):
                    candResumido = list(filter(lambda x: x.codigo == candAVerificar.codigo, listasCursos[pos][i]))[0]
                    candResumido.valido = "NAO"
                    ws = wb.get_sheet_by_name(f'Cota-{i}')
                    ws[f'D{listasCursos[pos][i].index(candResumido) + 2}'] = "NAO"
        wb.save('C:\\temp2\\' + arquivos[pos])
        wb.close()
        print("Candidato desclassificado na cota correspondente")


def arquivarChamadaPublica():
    for i in range(len(chamadasPublicas)):
        data = []
        for c in range(len(chamadasPublicas[i][2])):
            data.append([chamadasPublicas[i][2][c][0].codigo, chamadasPublicas[i][2][c][0].nome,
                         chamadasPublicas[i][2][c][0].inscricao, chamadasPublicas[i][2][c][0].tipoVaga,
                         chamadasPublicas[i][2][c][0].chamada, chamadasPublicas[i][2][c][1]])
        df = pd.DataFrame(data, columns=['Código', 'Nome', 'Inscrição', 'Tipo da Vaga', 'Chamada', "Matrícula"])
        df.to_excel(f'C:\\temp-ChamadaPublica\\CP-{listasCursos[i][0].nome[0]}-{listasCursos[i][0].nome[1]}.xlsx',
                    index=False)


def consolidarConferenciaPrincipal():
    arquivos = os.listdir('C:\\temp2')
    for i in range(len(arquivos)):
        wb = load_workbook('C:\\temp2\\' + arquivos[i])
        ws = wb.get_sheet_by_name('Cota-1')

        for j in range(len(listasCursos[i][1])):
            if any(listasCursos[i][1][j].codigo == cand[0].codigo and cand[0].chamada == "CP" for cand in chamadasPublicas[i][2]):
                ws[f'G{j+2}'] = "CP"

        wb.save('C:\\temp2\\' + arquivos[i])
        wb.close()