import math
from openpyxl import load_workbook
import os
import pandas as pd
import re

from Classes import *

cursos = []
pesos = [0, 10, 9, 8, 6, 7, 5, 4, 3, 1, 2]
percentuaisVagas = [.4, .00966, .03234, .04784, .16016, .00966, .03234, .04784, .16016, .05, .05]


def listarCandidatos():
    arquivos = os.listdir(('C:\\temp'))
    candidatos = []
    for arq in arquivos:
        nome = re.split(r'[-.]+', arq)
        cursos.append(Curso((nome[2], nome[3])))
        listaCandidatos = load_workbook('C:\\temp\\' + arq)
        for planilha in listaCandidatos:
            for linha in planilha:
                if (len(linha) > 1):
                    candidatos.append(Candidato(linha[5].value, linha[4].value, str(linha[2].value),
                                                str(linha[1].value), str(linha[3].value), linha[6].value,
                                                ajusteCotas(int(linha[6].value)), float(linha[12].value),
                                                int(linha[13].value)))
    return candidatos


def ajusteCotas(cotaInscricao):
    if cotaInscricao == 2:
        c = 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11
    elif cotaInscricao == 3:
        c = 1, 3, 5, 7, 9, 11
    elif cotaInscricao == 4:
        c = 1, 4, 5, 8, 9, 10
    elif cotaInscricao == 5:
        c = 1, 5, 9
    elif cotaInscricao == 6:
        c = 1, 6, 8, 9, 10, 11
    elif cotaInscricao == 7:
        c = 1, 7, 9, 11
    elif cotaInscricao == 8:
        c = 1, 8, 9, 10
    elif cotaInscricao == 9:
        c = 1, 9
    elif cotaInscricao == 10:
        c = 1, 10
    elif cotaInscricao == 11:
        c = 1, 11
    elif cotaInscricao == 12:
        c = 1, 2, 3, 4, 5, 6, 7, 8, 9, 10
    elif cotaInscricao == 13:
        c = 1, 3, 5, 7, 9
    elif cotaInscricao == 16:
        c = 1, 6, 8, 9, 10
    elif cotaInscricao == 17:
        c = 1, 7, 9
    else:
        c = 1,
    return c


def ordemChamada():
    ordem = []
    for num in range(11):
        ordem.append(Cota((num+1), percentuaisVagas[num], pesos[num]))
    return ordem


def fazerListasChamada(curso, candidatos, ordem):
    ordemCham = destinarVagas(ordem, curso)
    ordemCham.sort(key=lambda cota: cota.peso)
    curso.vagas = ordemCham

    onzeListas = []
    for i in range(1, 12, 1):
        onzeListas.append(listasParaCotas(curso, candidatos, i))
    dataCota1 = []
    for i in range(len(onzeListas[0])):
        dataCota1.append([onzeListas[0][i].codigo, onzeListas[0][i].nome, onzeListas[0][i].inscricao,
                          onzeListas[0][i].posicao, "SIM", None, 0, None, None, None, None, None])

    df = pd.DataFrame(dataCota1, columns=['Código', 'Nome', 'Inscrição', 'Posição', 'Válido', 'Matrícula', 'Chamada',
                                        'NC/desc RE', 'desc RI', 'desc PPI', 'desc EP', 'desc PCD'])
    df.to_excel(f'C:\\temp2\\{curso.nome[0]}-{curso.nome[1]}.xlsx', sheet_name=f'Cota-{1}', index=False)

    path = f'C:\\temp2\\{curso.nome[0]}-{curso.nome[1]}.xlsx'
    writer = pd.ExcelWriter(path, engine="openpyxl")
    writer.book = load_workbook(path)

    for i in range(1, 11):
        dataOutrasCotas = []
        for j in range(len(onzeListas[i])):
            dataOutrasCotas.append([onzeListas[i][j].codigo, onzeListas[i][j].nome, onzeListas[i][j].inscricao, "SIM"])
        df = pd.DataFrame(dataOutrasCotas, columns=['Código', 'Nome', 'Inscrição', 'Válido'])
        df.to_excel(writer, sheet_name=f'Cota-{i+1}', index=False)

    dataVagas = []
    for i in range(11):
        dataVagas.append([curso.vagas[i].cota, curso.vagas[i].vagas, 0, 0, 0])
    df = pd.DataFrame(dataVagas, columns=['Cota', 'VagasCh1', 'VagasCh2', 'VagasCh3', 'VagasChPubl'])
    df.to_excel(writer, sheet_name=f'Vagas', index=False)

    writer.save()
    writer.close()


def destinarVagas(ordem, curso):
    ordem.sort(key=lambda cota: cota.peso, reverse=True)

    vagas = 0
    while vagas == 0:
        try:
            vagas = int(input(f'Qual é o total de vagas para a chamada do curso {curso.nome[1]} - {curso.nome[0]}? '))
        except ValueError:
            print("Informação inválida.")

    if vagas < 1:
        print("Número de vagas inválidas")
    elif vagas == 1:
        ordem[10].vagas += 1
    else:
        vagasPublicas = math.ceil(vagas / 2)
        vagasUniversais = math.floor(vagas / 2)

        if vagasUniversais <= 3:
            ordem[10].vagas += 1
            vagasUniversais -= 1
            for i in range(8, 10, 1):
               if vagasUniversais > 0:
                    ordem[i].vagas += 1
                    vagasUniversais -= 1
        else:
            for i in range(8, 11):
                ordem[i].vagas = math.floor(vagasUniversais * ordem[i].vagas_perc * 2) \
                    if math.floor(vagasUniversais * ordem[i].vagas_perc*2) > 0 else 1
            for i in range(8, 11):
                vagasUniversais -= ordem[i].vagas
            if vagasUniversais > 0:
                for i in range(8, 11):
                    if vagasUniversais > 0:
                        ordem[i].vagas += 1
                        vagasUniversais -= 1
            if vagasUniversais < 0:
                for i in range(10, 7, -1):
                    if vagasUniversais < 0:
                        ordem[i].vagas -= 1
                        vagasUniversais += 1

        if vagasPublicas <= 8:
            sequencia = [9, 5, 7, 8, 3, 4, 6, 2]
            for i in range(8):
                for j in range(8):
                    if vagasPublicas > 0 and sequencia[i] == ordem[j].cota:
                        ordem[j].vagas += 1
                        vagasPublicas -= 1
        else:
            for i in range(8):
                ordem[i].vagas = math.floor(vagasPublicas * ordem[i].vagas_perc*2) \
                    if math.floor(vagasPublicas * ordem[i].vagas_perc*2) > 0 else 1
            for i in range(8):
                vagasPublicas -= ordem[i].vagas
            if vagasPublicas > 0:
                for i in range(8):
                    if vagasPublicas > 0:
                        ordem[i].vagas += 1
                        vagasPublicas -= 1
            if vagasPublicas < 0:
                for i in range(7, -1, -1):
                    if vagasPublicas < 0:
                        if ordem[i].vagas > 1:
                            ordem[i].vagas -= 1
                            vagasPublicas += 1
    return ordem


def listasParaCotas(curso, candidatos, cota):
    listaTemp = list(filter(lambda cand: cota in cand.cotas and cand.campus == curso.nome[1] and
                                         cand.curso == curso.nome[0], candidatos))
    listaTemp.sort(key=lambda cand: cand.posicao)
    return listaTemp